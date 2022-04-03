import json
import re

import openpyxl


def fun(file, shift):
    lenfile = "../../data/" + file + "/" + "len_" + file + ".txt"
    datalen = open(lenfile, encoding="utf-8").readlines()
    lenlist = list()
    for line in datalen:
        num = int(line.strip("\n"))
        lenlist.append(num)
    lenlist.sort()
    print(len(lenlist))
    starfile = "../../data/" + file + "/" + "star_height_" + file + ".txt"
    datastar = open(starfile, encoding="utf-8").readlines()
    starlist = list()
    for line in datastar:
        num = int(line.strip("\n"))
        starlist.append(num)
    print(len(starlist))
    starlist.sort()
    nestfile = "../../data/" + file + "/" + "nest_" + file + ".txt"
    datanest = open(nestfile, encoding="utf-8").readlines()
    nestlist = list()
    for line in datanest:
        num = int(line.strip("\n"))
        nestlist.append(num)
    print(len(nestlist))
    nestlist.sort()
    wb = openpyxl.load_workbook("../../data/data.xlsx")
    sheet = wb['Sheet1']  # 此时保证sheet1存在

    lenth = len(lenlist)
    # s = int(lenth/10)
    # e = int(lenth/10*9)
    s = 0
    e = lenth
    index = 1
    for i in range(s, e):
        sheet.cell(index + 1, 1 + shift * 3).value = lenlist[i]
        sheet.cell(index + 1, 2 + shift * 3).value = starlist[i]
        sheet.cell(index + 1, 3 + shift * 3).value = nestlist[i]
        index += 1
    wb.save("../../data/data.xlsx")


def funjs(file, shift):
    lenfile = "../../data/" + file + "/" + "len_" + file + ".txt"
    datalen = open(lenfile, encoding="utf-8").readlines()
    lenlist = list()
    for line in datalen:
        num = int(line.strip("\n"))
        lenlist.append(num)
    lenlist.sort()
    print(len(lenlist))
    starfile = "../../data/" + file + "/" + "star_height_" + file + ".txt"
    datastar = open(starfile, encoding="utf-8").readlines()
    starlist = list()
    for line in datastar:
        num = int(line.strip("\n"))
        starlist.append(num)
    print(len(starlist))
    starlist.sort()
    nestfile = "../../data/" + file + "/" + "nest_" + file + ".txt"
    datanest = open(nestfile, encoding="utf-8").readlines()
    nestlist = list()
    for line in datanest:
        num = int(line.strip("\n"))
        nestlist.append(num)
    print(len(nestlist))
    nestlist.sort()
    wb = openpyxl.load_workbook("../../data/data.xlsx")
    sheet = wb['Sheet1']  # 此时保证sheet1存在

    lenth = len(lenlist)/20
    # s = int(lenth/10)
    # e = int(lenth/10*9)
    s = 0
    e = int(lenth)
    index = 1
    for i in range(s, e):
        sheet.cell(index + 1, 1 + shift * 3).value = lenlist[i*20]
        sheet.cell(index + 1, 2 + shift * 3).value = starlist[i*20]
        sheet.cell(index + 1, 3 + shift * 3).value = nestlist[i*20]
        index += 1
    wb.save("../../data/data.xlsx")
fun("java", 0)
fun("python",1)
# fun("js",2)
fun("csharp",3)
fun("php",4)
fun("perl",5)

# fun("php")
